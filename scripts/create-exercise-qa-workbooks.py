from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREY_FILL = PatternFill("solid", fgColor="F7F8FA")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")

MAIN_HEADERS = [
    "Exercise ID", "Subject", "Topic", "Type", "Difficulty", "Question",
    "Answer 1", "Answer 2", "Answer 3", "Answer 4", "Correct answer / solution",
    "Hint 1", "Hint 2", "LP21 fit", "Correctness", "Language", "Reviewer notes",
    "Decision", "Reviewed by", "Reviewed date", "Report status", "Reported at",
    "Report reason", "Correction made", "Fixed at", "Retest URL",
]

SPECIAL_TYPES = {"counting", "matching", "memory", "drag-drop", "number-line", "word-search"}


def style_header(ws, width: int) -> None:
    for cell in ws[1][:width]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34


def add_dropdown(ws, column: str, values: list[str], end_row: int) -> None:
    validation = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    validation.error = "Bitte einen Wert aus der Liste wählen."
    validation.errorTitle = "Ungültiger Wert"
    ws.add_data_validation(validation)
    validation.add(f"{column}2:{column}{end_row}")


def add_table(ws, name: str, end_column: str, end_row: int) -> None:
    table = Table(displayName=name, ref=f"A1:{end_column}{end_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def create_workbook(source: Path, output: Path) -> dict:
    data = json.loads(source.read_text())
    grade = data["grade"]
    rows = data["rows"]
    topics = data["topics"]

    wb = Workbook()
    main = wb.active
    main.title = f"Grade {grade} QA"
    main.append(MAIN_HEADERS)

    for item in rows:
        options = (item["options"] + [None] * 4)[:4]
        hints = (item["hints"] + [None] * 2)[:2]
        main.append([
            item["exerciseId"], item["subject"], item["topic"], item["type"], item["difficulty"], item["question"],
            *options, item["correct"], *hints, None, None, None, None, None, None, None,
            item.get("reportStatus"), item.get("reportedAt"), item.get("reportReason"),
            item.get("correctionMade"), item.get("fixedAt"), item.get("retestUrl"),
        ])
        row_number = main.max_row
        if item["type"] == "multiple-choice" and item["storedAnswer"] in item["options"][:4]:
            answer_column = 7 + item["options"][:4].index(item["storedAnswer"])
        else:
            answer_column = 11
        main.cell(row_number, answer_column).fill = GREEN_FILL

    style_header(main, len(MAIN_HEADERS))
    main.freeze_panes = "D2"
    main.auto_filter.ref = f"A1:Z{main.max_row}"
    main.sheet_view.showGridLines = False
    add_table(main, f"Grade{grade}QATable", "Z", main.max_row)
    add_dropdown(main, "N", ["Yes", "Questionable", "No"], main.max_row)
    add_dropdown(main, "O", ["Correct", "Wrong", "Unclear"], main.max_row)
    add_dropdown(main, "P", ["Good", "Needs work"], main.max_row)
    add_dropdown(main, "R", ["Approved", "Fix", "Remove"], main.max_row)
    add_dropdown(main, "U", ["new", "reviewed", "fixed", "wont_fix", "retest_passed"], main.max_row)
    widths = {"A": 18, "B": 28, "C": 30, "D": 18, "E": 11, "F": 65, "G": 28, "H": 28, "I": 28, "J": 28,
              "K": 42, "L": 48, "M": 48, "N": 15, "O": 15, "P": 16, "Q": 50, "R": 14, "S": 18, "T": 16,
              "U": 16, "V": 22, "W": 54, "X": 62, "Y": 18, "Z": 56}
    for column, width in widths.items():
        main.column_dimensions[column].width = width
    for row in main.iter_rows(min_row=2, max_row=main.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    special = wb.create_sheet("Special exercises")
    special_headers = ["Exercise ID", "Subject", "Topic", "Type", "Difficulty", "Question", "Readable solution", "Technical structure", "Hint 1", "Hint 2"]
    special.append(special_headers)
    special_rows = [item for item in rows if item["type"] in SPECIAL_TYPES]
    for item in special_rows:
        hints = (item["hints"] + [None] * 2)[:2]
        special.append([item["exerciseId"], item["subject"], item["topic"], item["type"], item["difficulty"], item["question"], item["correct"], item["structure"], *hints])
        special.cell(special.max_row, 7).fill = GREEN_FILL
    style_header(special, len(special_headers))
    special.freeze_panes = "D2"
    special.auto_filter.ref = f"A1:J{max(1, special.max_row)}"
    special.sheet_view.showGridLines = False
    if special.max_row > 1:
        add_table(special, f"Grade{grade}SpecialTable", "J", special.max_row)
    for column, width in {"A": 18, "B": 28, "C": 30, "D": 18, "E": 11, "F": 65, "G": 55, "H": 80, "I": 48, "J": 48}.items():
        special.column_dimensions[column].width = width
    for row in special.iter_rows(min_row=2, max_row=special.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    progress = wb.create_sheet("Progress")
    progress.append([f"Klasse {grade} QA progress"])
    progress["A1"].font = Font(size=18, bold=True, color="1F4E78")
    progress.append(["Total exercises", len(rows)])
    progress.append(["Reviewed", f'=COUNTIF(\'{main.title}\'!R2:R{main.max_row},"<>")'])
    progress.append(["Approved", f'=COUNTIF(\'{main.title}\'!R2:R{main.max_row},"Approved")'])
    progress.append(["Fix", f'=COUNTIF(\'{main.title}\'!R2:R{main.max_row},"Fix")'])
    progress.append(["Remove", f'=COUNTIF(\'{main.title}\'!R2:R{main.max_row},"Remove")'])
    progress.append(["Remaining", "=B2-B3"])
    progress.append(["Reviewed %", "=IFERROR(B3/B2,0)"])
    progress["B8"].number_format = "0.0%"
    progress.append([])
    progress.append(["Subject", "Topic", "Total", "Reviewed", "Approved", "Fix", "Remove", "Remaining", "Reviewed %"])
    start = 11
    for item in topics:
        progress.append([
            item["subject"], item["topic"], item["count"],
            f'=COUNTIFS(\'{main.title}\'!B:B,A{progress.max_row + 1},\'{main.title}\'!C:C,B{progress.max_row + 1},\'{main.title}\'!R:R,"<>")',
            f'=COUNTIFS(\'{main.title}\'!B:B,A{progress.max_row + 1},\'{main.title}\'!C:C,B{progress.max_row + 1},\'{main.title}\'!R:R,"Approved")',
            f'=COUNTIFS(\'{main.title}\'!B:B,A{progress.max_row + 1},\'{main.title}\'!C:C,B{progress.max_row + 1},\'{main.title}\'!R:R,"Fix")',
            f'=COUNTIFS(\'{main.title}\'!B:B,A{progress.max_row + 1},\'{main.title}\'!C:C,B{progress.max_row + 1},\'{main.title}\'!R:R,"Remove")',
            f"=C{progress.max_row + 1}-D{progress.max_row + 1}",
            f"=IFERROR(D{progress.max_row + 1}/C{progress.max_row + 1},0)",
        ])
    for cell in progress[10]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    progress.freeze_panes = "A11"
    progress.sheet_view.showGridLines = False
    progress.column_dimensions["A"].width = 30
    progress.column_dimensions["B"].width = 34
    for column in "CDEFGHI":
        progress.column_dimensions[column].width = 14
    for row in range(start, progress.max_row + 1):
        progress.cell(row, 9).number_format = "0.0%"
    progress.conditional_formatting.add(f"F{start}:F{progress.max_row}", FormulaRule(formula=[f"F{start}>0"], fill=YELLOW_FILL))
    progress.conditional_formatting.add(f"G{start}:G{progress.max_row}", FormulaRule(formula=[f"G{start}>0"], fill=RED_FILL))

    wb.save(output)
    return {
        "grade": grade,
        "exercises": len(rows),
        "topics": len(topics),
        "special": len(special_rows),
        "duplicateIds": len(data["duplicateIds"]),
        "output": str(output),
        "types": dict(Counter(item["type"] for item in rows)),
        "subjects": dict(Counter(item["subject"] for item in rows)),
    }


def main() -> None:
    grades = [int(value) for value in sys.argv[1:]] or [2, 3, 4, 5, 6]
    for grade in grades:
        result = create_workbook(Path(f"/tmp/cleverli-grade-{grade}-qa.json"), Path(f"/tmp/Cleverli-Klasse-{grade}-Exercise-QA.xlsx"))
        print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
